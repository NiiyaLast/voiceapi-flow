import pandas as pd
import os
# from .structured_data import process_excel_with_ai 
from openpyxl import load_workbook, Workbook
import subprocess
import platform

def export_to_excel(results, filename):
    """
    将识别结果导出为 Excel 文件。支持单个数据集和多个数据集（对象数组）。

    :param results: 
        - 单个数据集：包含识别结果的列表，每个元素是一个字典，格式为 {"time": 时间, "result": 结果}
        - 多个数据集：对象数组，每个对象包含数据和sheet名称，格式为 [{"data": [...], "sheet_name": "..."}, ...]
    :param filename: 导出的 Excel 文件名
    """
  
    output_dir = "./download"
    os.makedirs(output_dir, exist_ok=True)

    # 拼接完整路径
    filepath = os.path.join(output_dir, filename)

    # 检查是否为对象数组（多个数据集）
    if isinstance(results, list) and len(results) > 0 and isinstance(results[0], dict):
        # 处理对象数组，每个对象写入一个sheet
        export_multiple_sheets_to_excel(results, filepath)
    else:
        # 处理单个数据集
        export_single_sheet_to_excel(results, filepath)

    print(f"Results exported to {filepath}")
    open_excel_file(filepath)


def export_single_sheet_to_excel(results, filepath):
    """
    将单个数据集导出为 Excel 文件（单个sheet）。

    :param results: 包含识别结果的列表，每个元素是一个字典
    :param filepath: 导出的 Excel 文件完整路径
    """
    # 将结果转换为 pandas DataFrame
    df = pd.DataFrame(results)

    # 导出为 Excel 文件
    df.to_excel(filepath, index=False)


def export_multiple_sheets_to_excel(results_array, filepath):
    """
    将多个数据集导出为 Excel 文件（多个sheet）。

    :param results_array: 对象数组，每个对象包含数据和sheet名称
                         格式为 [{"data": [...], "sheet_name": "..."}, ...]
    :param filepath: 导出的 Excel 文件完整路径
    """
    try:
        # 创建 ExcelWriter 对象
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            for i, result_obj in enumerate(results_array):
                # 获取数据和sheet名称
                data = result_obj.get('data', [])
                sheet_name = result_obj.get('name', f'Sheet{i+1}')
                
                # 确保sheet名称符合Excel命名规范
                sheet_name = sanitize_sheet_name(sheet_name)
                
                # 将数据转换为DataFrame
                if data:  # 如果有数据
                    df = pd.DataFrame(data)
                    # 写入指定的sheet
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"数据已写入sheet: {sheet_name}，共 {len(data)} 行数据")
                else:
                    # 如果没有数据，创建空的sheet
                    pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"创建空sheet: {sheet_name}")
        
        print(f"成功创建包含 {len(results_array)} 个sheet的Excel文件")
        
    except Exception as e:
        print(f"创建多sheet Excel文件时出错: {e}")
        # 回退到单sheet模式
        all_data = []
        for result_obj in results_array:
            data = result_obj.get('data', [])
            all_data.extend(data)
        export_single_sheet_to_excel(all_data, filepath)


def sanitize_sheet_name(name):
    """
    清理sheet名称，确保符合Excel命名规范。
    
    :param name: 原始sheet名称
    :return: 清理后的sheet名称
    """
    # Excel sheet名称不能包含这些字符: \ / * ? [ ] :
    invalid_chars = ['\\', '/', '*', '?', '[', ']', ':']
    sanitized = name
    
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '_')
    
    # Excel sheet名称最大长度为31个字符
    if len(sanitized) > 31:
        sanitized = sanitized[:31]
    
    # 不能为空或只包含空格
    if not sanitized or sanitized.isspace():
        sanitized = "Sheet1"
    
    return sanitized


def export_to_excel_sheetn(results, filename, sheet_name="Sheet2"):
    """
    将数据写入指定Excel文件的Sheet2（或指定工作表）中
    
    :param results: 包含数据的列表，每个元素是一个字典
    :param target_excel_path: 目标Excel文件路径
    :param sheet_name: 工作表名称，默认为"Sheet2"
    """

    output_dir = "./download"
    os.makedirs(output_dir, exist_ok=True)

    # 拼接完整路径
    filepath = os.path.join(output_dir, filename)

    try:
        # 检查目标Excel文件是否存在
        if os.path.exists(filepath):
            # 加载现有工作簿
            workbook = load_workbook(filepath)
            print(f"加载现有Excel文件: {filepath}")
        else:
            # 创建新工作簿
            workbook = Workbook()
            print(f"创建新Excel文件: {filepath}")
        
        # 检查是否存在指定的工作表
        if sheet_name in workbook.sheetnames:
            # 删除现有的工作表（清空数据）
            del workbook[sheet_name]
            print(f"删除现有工作表: {sheet_name}")
        
        # 创建新的工作表
        worksheet = workbook.create_sheet(title=sheet_name)
        
        # 将数据转换为DataFrame
        df = pd.DataFrame(results)
        
        # 写入表头
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_idx, value=header)
        
        # 写入数据
        for row_idx, row_data in enumerate(df.values, 2):  # 从第2行开始写入数据
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # 保存工作簿
        workbook.save(filepath)
        print(f"数据已成功写入 {filepath} 的 {sheet_name} 工作表")
        print(f"共写入 {len(results)} 行数据，{len(headers)} 列")
        
        return True
        
    except Exception as e:
        print(f"写入Excel文件时出错: {e}")
        return False
    
def open_excel_file(filepath):
    """
    跨平台打开 Excel 文件
    
    :param filepath: Excel 文件的完整路径
    """
    try:
        # 获取绝对路径
        abs_filepath = os.path.abspath(filepath)
        
        # 根据操作系统选择合适的打开方式
        system = platform.system()
        
        if system == "Windows":
            # Windows 系统
            os.startfile(abs_filepath)
            print(f"已在 Windows 中打开文件: {abs_filepath}")
            
        elif system == "Darwin":  # macOS
            # macOS 系统
            subprocess.run(["open", abs_filepath])
            print(f"已在 macOS 中打开文件: {abs_filepath}")
            
        elif system == "Linux":
            # Linux 系统
            subprocess.run(["xdg-open", abs_filepath])
            print(f"已在 Linux 中打开文件: {abs_filepath}")
            
        else:
            print(f"不支持的操作系统: {system}")
            print(f"请手动打开文件: {abs_filepath}")
            
    except Exception as e:
        print(f"打开文件时出错: {e}")
        print(f"请手动打开文件: {filepath}")